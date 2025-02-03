import os
import glob
import linecache
import pandas as pd
import openpyxl

if __name__ == '__main__':
   
    f = openpyxl.load_workbook(r'D:\\VTM-er\\0919_nnvc-6.0_libtorch-1.11\\1031_test\\3_txt\\1_extract.xlsx')
    f1 = f['Sheet1']
    count = 0
    folder_name = "D:\\VTM-er\\0919_nnvc-6.0_libtorch-1.11\\1031_test\\3_txt\\"  # 获取文件夹的名字，即路径
    file_names = os.listdir(folder_name)  # 获取文件夹内所有文件的名字
    i=0
    for file_name in file_names:
        if file_name.endswith('.txt'):
            print(file_name)
            f1.cell(row=i + 1, column=1).value = os.path.splitext(file_name)[0]
            b=0
            with open(folder_name+file_name, 'r') as file:
                for line_number, line in enumerate(file):
                    if 'POC' in line:
                        a=line_number
                        b=a-1
                        print('b:',b)
                        #print(f'Line {line_number-1}
                print(linecache.getline(folder_name+file_name, b+1))
                print(linecache.getline(folder_name + file_name, b + 6))


                f1.cell(row=i + 1, column=3).value = linecache.getline(folder_name+file_name, b+1)
                f1.cell(row=i + 1, column=2).value = linecache.getline(folder_name+file_name, b+6)
            count += 1
            print(count)
            i+=1
            f.save(r'D:\\VTM-er\\0919_nnvc-6.0_libtorch-1.11\\1031_test\\3_txt\\1_extract.xlsx')
    f.save(r'D:\\VTM-er\\0919_nnvc-6.0_libtorch-1.11\\1031_test\\3_txt\\1_extract.xlsx')
    print("提取完成！")
